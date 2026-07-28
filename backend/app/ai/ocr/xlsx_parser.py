# =============================================================================
# File: xlsx_parser.py
# Module/Service: Pipeline Worker — Document Parsing & Cleaning ([AI])
# Layer: Service
# Purpose: XLSX sheet extraction as semantic row/column text.
# Responsibilities:
#   - Collect sheet rows; guess header row; emit table blocks per sheet
# Dependencies:
#   - openpyxl; app.ai.ocr.constants/models/tables
# Public Exports:
#   - _parse_xlsx, _xlsx_collect_rows, _xlsx_rows_to_semantic,
#     _xlsx_guess_header_row, _looks_numeric
# Database/Table: N/A
# Related Modules: app.ai.ocr.*
# Important Notes: page_count = sheet count; page_number = sheet index.
# =============================================================================

from __future__ import annotations

import io

from .constants import XLSX_MAX_HEADER_SCAN
from .models import _ParsedBlock
from .tables import _format_kv_fallback, _format_table_semantic, _infer_table_col_count


def _parse_xlsx(data: bytes) -> tuple[list[_ParsedBlock], int]:
    """Extract XLSX sheets as semantic row/column text (not pipe-joined)."""
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to open XLSX: {exc}") from exc

    blocks: list[_ParsedBlock] = []
    try:
        for idx, sheet in enumerate(wb.worksheets, start=1):
            rows = _xlsx_collect_rows(sheet)
            if not rows:
                continue
            text = _xlsx_rows_to_semantic(rows, sheet_title=sheet.title)
            if text.strip():
                blocks.append(
                    _ParsedBlock(
                        text=text,
                        page_number=idx,
                        section=sheet.title,
                        block_type="table",
                        table_col_count=_infer_table_col_count(text),
                    )
                )
        return blocks, len(wb.worksheets)
    finally:
        wb.close()


def _xlsx_collect_rows(sheet: object) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
        cells = ["" if c is None else str(c).strip() for c in row]
        # Trim trailing empties
        while cells and not cells[-1]:
            cells.pop()
        if any(cells):
            rows.append(cells)
    return rows


def _xlsx_rows_to_semantic(rows: list[list[str]], *, sheet_title: str) -> str:
    """Convert sheet rows to ``Row N / Col = Val`` semantic text."""
    if not rows:
        return ""

    header_idx = _xlsx_guess_header_row(rows)
    if header_idx is None:
        return _format_kv_fallback(rows)

    headers = rows[header_idx]
    body = rows[header_idx + 1 :]
    return _format_table_semantic(headers, body, title=f"Sheet: {sheet_title}")


def _xlsx_guess_header_row(rows: list[list[str]]) -> int | None:
    """Pick the first non-empty row with mostly non-numeric labels as header."""
    limit = min(len(rows), XLSX_MAX_HEADER_SCAN)
    for i in range(limit):
        row = rows[i]
        non_empty = [c for c in row if c]
        if len(non_empty) < 1:
            continue
        numericish = sum(1 for c in non_empty if _looks_numeric(c))
        if numericish / len(non_empty) <= 0.4:
            return i
    return 0 if rows else None


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False
